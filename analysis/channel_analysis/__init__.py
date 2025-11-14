import copy
import itertools
import os
from collections import defaultdict
from typing import List, Set, Dict

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from analysis.channel_analysis.file_management import download_remote_configs, download_remote_results
from net.DL_config import Config, get_base_config
from utility.paths import get_path_config, get_path_results
from utility.stats import Results


def count_selected_channels_across_folds(base_dir, configs: List[Config], output_path: str = None):
    for config in configs:
        config_path = os.path.join(base_dir, "..", "..", get_path_config(config, config.get_name()))
        if os.path.exists(config_path):
            config.load_config(config_path, config.get_name())
        else:
            print(f"Config not found for {config.get_name()}")
            continue
        nb_folds = config.nb_folds
        channels = config.selected_channels
        channel_count = {}
        for fold in channels:
            for ch in channels[fold]:
                if ch not in channel_count:
                    channel_count[ch] = 1
                else:
                    channel_count[ch] += 1

        sorted_channels = sorted(channel_count.items(), key=lambda x: x[1], reverse=True)
        print(f"Number of folds a channel is chosen: {sorted_channels}")
        # make a hbar plot with channels on y-axis and counts on x-axis + sort them
        channels, counts = zip(*sorted_channels)

        # Create horizontal bar plot
        plt.figure(figsize=(10, 6))
        plt.barh(channels, counts, color='skyblue')
        plt.xlabel("Number of folds the channel is selected out of the {} folds".format(nb_folds))
        plt.title("Channel Counts")
        plt.gca().invert_yaxis()  # Invert y-axis to have the highest count on top

        if output_path:
            output_file = os.path.join(output_path, "channel_counts", f"{config.get_name()}_channel_counts.png")
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            plt.savefig(output_file)
        else:
            plt.show()

        plt.close()

def mine_frequent_channels(base_dir, configs: List[Config], output_path: str = None, min_support: float = 0.5):
    from mlxtend.frequent_patterns import apriori, association_rules
    from mlxtend.preprocessing import TransactionEncoder
    for config in configs:
        config_path = os.path.join(base_dir, "..", "..", get_path_config(config, config.get_name()))
        if os.path.exists(config_path):
            config.load_config(config_path, config.get_name())
        else:
            print(f"Config not found for {config.get_name()}")
            continue
        channels = config.selected_channels
        transactions = []
        for fold in channels:
            transactions.append(list(channels[fold]))

        # Convert transactions to a one-hot encoded DataFrame
        te = TransactionEncoder()
        te_ary = te.fit(transactions).transform(transactions)
        df = pd.DataFrame(te_ary, columns=te.columns_)

        # Apply Apriori algorithm
        frequent_itemsets = apriori(df, min_support=min_support, use_colnames=True)
        rules = association_rules(frequent_itemsets, metric="support", min_threshold=min_support)

        print(f"Frequent itemsets for {config.get_name()}:")
        print(frequent_itemsets)
        print(f"Association rules for {config.get_name()}:")
        print(rules)

        if output_path:
            output_file = os.path.join(output_path, "frequent_channels", f"{config.get_name()}_frequent_channels.csv")
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            # Save the frequent itemsets and rules to same CSV file
            frequent_itemsets.to_csv(output_file, index=False)
            with open(output_file, 'a') as f:
                f.write('\n')
            rules.to_csv(output_file, index=False, mode='a')

def find_interchangeable_channels(base_dir, configs: List[Config], output_path:str, minimal_support: float = 0.5):
    assert 0 <= minimal_support <= 1
    for i, config in enumerate(configs):
        config_path = os.path.join(base_dir, "..", "..", get_path_config(config, config.get_name()))
        if os.path.exists(config_path):
            config.load_config(config_path, config.get_name())
        else:
            print(f"Config not found for {config.get_name()}")
            continue

        config.nb_folds = len(config.folds)
        # results = []
        # results_path = os.path.join(base_dir, "..", "..", get_path_results(config, config.get_name()))
        # results.append(Results(config))
        # if os.path.exists(results_path):
        #     results[i].load_results(results_path)
        # else:
        #     print(f"Results not found for {config.get_name()}")
        #     continue
        #
        # result = results[i]
        included_channels = sorted(config.included_channels)
        # TODO: possibly change these two lines
        selected_channels = [config.selected_channels[fold_i] for fold_i in range(config.nb_folds)]
        clusters = [config.channel_selector[fold_i].clusters for fold_i in range(config.nb_folds)]
        # Count how many times two channels appears in the same cluster
        co_occurrence_matrix = np.zeros((len(included_channels), len(included_channels)))
        nb_folds_channel_selected = np.zeros((len(included_channels), len(included_channels)), dtype=int)
        for clustering_i in clusters:
            for cluster in clustering_i:
                for ch1, ch2 in itertools.combinations(cluster, 2):
                    co_occurrence_matrix[ch1, ch2] += 1
                    co_occurrence_matrix[ch2, ch1] += 1

                for ch in cluster:
                    co_occurrence_matrix[ch, ch] += 1
                    nb_folds_channel_selected[ch, ch] += 1

            channels_in_all_clusters = list(itertools.chain.from_iterable(clustering_i))
            for ch1, ch2 in itertools.combinations(channels_in_all_clusters, 2):
                nb_folds_channel_selected[ch1, ch2] += 1
                nb_folds_channel_selected[ch2, ch1] += 1

        # Normalize the co-occurrence matrix by the  number of folds each channel got past the irrelevant selector
        for m in range(co_occurrence_matrix.shape[0]):
            for n in range(m, co_occurrence_matrix.shape[1]):
                co_occurrence_matrix[m, n] /= nb_folds_channel_selected[m, n]
                if m != n:
                    co_occurrence_matrix[n, m] /= nb_folds_channel_selected[m, n]
                    assert co_occurrence_matrix[m, n] == co_occurrence_matrix[n, m] or \
                    (np.isnan(co_occurrence_matrix[m, n]) and np.isnan(co_occurrence_matrix[n, m]))


        # Average the evaluation metric for channel selection across the folds to get an indication of how relevant
        # the channel is
        average_relevance_channels = {ch: 0 for ch in range(len(included_channels))}
        for fold_i in range(config.nb_folds):
            evaluation_metrics = config.channel_selector[fold_i].evaluation_metric_per_channel
            evaluation_metrics.update(config.channel_selector[fold_i].removed_series_too_low_metric)
            for ch, metric in evaluation_metrics.items():
                average_relevance_channels[ch] += metric
        # Average the relevance across folds
        average_relevance_channels = {ch: metric / config.nb_folds for ch, metric in average_relevance_channels.items()}

        # Select all channels that have a co-occurrence above a certain threshold
        processed_co_occurrence = process_co_occurrence_matrix(co_occurrence_matrix, minimal_support)

        # Add relevance to results
        def co_occurrence2str(chs: dict) -> Dict[str, dict | str]:
            result = {}
            for key, value in chs.items():
                chs_relevance = {included_channels[c]: average_relevance_channels[c] for c in key}
                sorted_chs = [f"{k} ({'%.2f' % v})" for k, v in sorted(chs_relevance.items(), key=lambda item: item[1], reverse=True)]
                key_str = ', '.join(sorted_chs)

                if value == []:
                    value_str = ""
                else:
                    value_str = co_occurrence2str(value)

                result[key_str] = value_str
            return result

        processed_co_occurrence_with_relevance = co_occurrence2str(processed_co_occurrence)


        # for key, value in processed_co_occurrence.items():
        #
        # for c in processed_co_occurrence:
        #     c_relevance = {included_channels[ch]: average_relevance_channels[ch] for ch in c}
        #     sorted_c = [f"{k}: {'%.2f' % v}" for k, v in sorted(c_relevance.items(), key=lambda item: item[1], reverse=True)]
        #     processed_co_occurrence_with_relevance.append(sorted_c)

        # Save selected channels, co-occurrence matrix and average_relevance to a CSV file
        output_file = os.path.join(output_path, f"{config.get_name()}_interchangeable_channels.xlsx")
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            workbook = writer.book
            sheet = workbook.create_sheet("Interchangeable Channels")
            writer.sheets["Interchangeable Channels"] = sheet

            # Title font and alignment
            title_font = Font(bold=True, size=12)
            alignment = Alignment(horizontal="center")

            # Write title and selected channels
            sheet.cell(row=1, column=1, value="Selected Channels").font = title_font
            sheet.cell(row=1, column=1).alignment = alignment

            selected_channels_df = pd.DataFrame(
                selected_channels,
                index=[f"Fold {i}" for i in config.selected_channels.keys()]
            )
            selected_channels_df.to_excel(writer, sheet_name="Interchangeable Channels", startrow=2)

            # Adjust column widths for selected channels
            for col_num in range(1, len(selected_channels_df.columns) + 1):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = 15

            # Write title and processed co-occurrence
            start_row = 2 + len(selected_channels_df) + 4  # Add extra whitespace
            sheet.cell(row=start_row, column=1, value="Processed Co-occurrence with Relevance").font = title_font
            sheet.cell(row=start_row, column=1).alignment = alignment

            start_row += 1
            # Write each list in processed_co_occurrence_with_relevance as a row
            def write_co_occurrence_to_sheet(co_occurrence_dict, sheet, start_row, start_column):
                row_nb = start_row
                for key, value in co_occurrence_dict.items():
                    splitted_key = key.split(", ")
                    for j, v in enumerate(splitted_key, start=start_column):
                        sheet.cell(row=row_nb, column=j, value=v)
                    row_nb += 1
                    if isinstance(value, dict):
                        # If value is a dict, write it recursively
                        extra_rows = write_co_occurrence_to_sheet(value, sheet, row_nb, start_column+1)
                        row_nb += extra_rows
                return row_nb - start_row
            start_row += write_co_occurrence_to_sheet(processed_co_occurrence_with_relevance, sheet, start_row, 1)

            # Write title and co-occurrence matrix
            # start_row += len(processed_co_occurrence_with_relevance) + 4  # Add extra whitespace
            start_row += 4  # Add extra whitespace
            sheet.cell(row=start_row, column=1, value="Co-occurrence Matrix").font = title_font
            sheet.cell(row=start_row, column=1).alignment = alignment

            co_occurrence_df = pd.DataFrame(
                co_occurrence_matrix,
                index=included_channels,
                columns=included_channels
            )
            co_occurrence_df.to_excel(writer, sheet_name="Interchangeable Channels", startrow=start_row + 1)

            # Write title and average relevance
            start_row += len(co_occurrence_df) + 4  # Add extra whitespace
            sheet.cell(row=start_row, column=1, value="Average Relevance").font = title_font
            sheet.cell(row=start_row, column=1).alignment = alignment

            average_relevance_df = pd.DataFrame(
                list(average_relevance_channels.values()),
                columns=["Average Relevance"],
                index=included_channels
            )
            average_relevance_df.to_excel(writer, sheet_name="Interchangeable Channels", startrow=start_row + 1)

            # Adjust column widths for all dataframes
            for col_num in range(1, average_relevance_df.shape[1] + 1):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = 20

        print(f"Interchangeable channels for {config.get_name()} saved to {output_file}")
        # Combine the count and averaged evaluation metric to find the channels that are equally relevant, but redundant


def process_co_occurrence_matrix(co_occurrence: np.ndarray, minimal_support=0.5):
    assert 0 <= minimal_support <= 1

    groups: List[Set] = []

    for i in range(co_occurrence.shape[0]):
        # Find all indices j where co_occurrence[i, j] >= minimal_support
        # and j >= i to avoid duplicates
        group = {i}
        for j in range(i + 1, co_occurrence.shape[1]):
            if co_occurrence[i, j] >= minimal_support:
                group.add(j)

        to_remove = set()
        for m, n in itertools.combinations(group, 2):
            if co_occurrence[m, n] < minimal_support:
                to_remove.add(n)
        group.difference_update(to_remove)

        if not all(np.isnan(co_occurrence[i])):
            # Check if the group is already in groups
            if not any(group.issubset(existing_group) for existing_group in groups):
                # If not, add it to groups
                groups.append(group)

    # index_structure = defaultdict(list)
    # for i, group in enumerate(groups):
    #     for elem in group:
    #         index_structure[elem].append(i)

    def list_to_tree(sets: List[Set]):
        result = dict()
        sets = sorted(sets, key=len)
        already_handled = set()
        for s_i, s in enumerate(sets):
            if s_i in already_handled:
                continue
            common = [sets[k] for k in range(s_i + 1, len(sets)) if len(s.intersection(sets[k])) > 0]
            if not common:
                result[frozenset(s)] = []
            else:
                intersection = set.intersection(*common, s)
                if not intersection:
                    result[frozenset(s)] = []
                    continue
                differences = [other.difference(intersection) for other in [s, *common]]
                to_remove_differences = set()
                for diff_i, diff in enumerate(differences):
                    for elem in diff:
                        others = set.union(*[d for d_i, d in enumerate(differences) if d_i != diff_i])
                        values = [co_occurrence[elem, o] for o in others if (co_occurrence[elem, o] < minimal_support or
                                  np.isnan(co_occurrence[elem, o]))]
                        if all(np.isnan(values)):
                            intersection.add(elem)
                            to_remove_differences.add(elem)

                differences = [{d for d in diff if d not in to_remove_differences} for diff in differences]
                differences = [d for d in differences if len(d) > 0]
                if not differences:
                    result[frozenset(intersection)] = []
                else:
                    result[frozenset(intersection)] = list_to_tree(differences)
                already_handled.update({sets.index(c) for c in common})

        return result

    return list_to_tree(groups)

def construct_set_selected_channels(base_dir, configs: List[Config], output_path: str, minimal_support: float = 0.2,
                                    threshold_metric: float = 0.5, metric:str = 'score'):
    assert 0 <= minimal_support <= 1
    for i, config in enumerate(configs):
        config_save_dir = config.save_dir
        config_path = os.path.join(base_dir, "..", "..", get_path_config(config, config.get_name()))
        results_path = os.path.join(base_dir, "..", "..", get_path_results(config, config.get_name()))
        base_config = copy.deepcopy(config)
        base_config.channel_selection = False
        base_config.add_to_name = ""
        base_config_path = os.path.join(base_dir, "..", "..", get_path_config(base_config, base_config.get_name()))
        base_results_path = os.path.join(base_dir, "..", "..", get_path_results(base_config, base_config.get_name()))
        if os.path.exists(config_path):
            config.load_config(config_path, config.get_name())
        else:
            print(f"Config not found for {config.get_name()}")
            continue

        if os.path.exists(base_config_path):
            base_config.load_config(base_config_path, base_config.get_name())
        else:
            download_remote_configs([base_config], local_base_dir=config_save_dir)
            if os.path.exists(base_config_path):
                base_config.load_config(base_config_path, base_config.get_name())
            else:
                print(f"Base config not found for {base_config.get_name()}")
                continue

        config.nb_folds = len(config.folds)
        results = Results(config)
        if os.path.exists(results_path):
            results.load_results(results_path)
        else:
            print(f"Results not found for {config.get_name()}")
            continue

        base_results = Results(base_config)
        if os.path.exists(base_results_path):
            base_results.load_results(base_results_path)
        else:
            download_remote_results([base_config], local_base_dir=config_save_dir)
            if os.path.exists(base_results_path):
                base_results.load_results(base_results_path)
            else:
                print(f"Base results not found for {base_config.get_name()}")
                continue

        selected_channels = [config.selected_channels[fold_i] for fold_i in range(config.nb_folds)]
        selected_channels_counts = defaultdict(int)
        for fold_i in range(config.nb_folds):
            for ch in config.selected_channels[fold_i]:
                selected_channels_counts[ch] += 1
        all_selected_channels = set(selected_channels_counts.keys())
        channel_combinations = []
        for r in range(1, len(all_selected_channels) + 1):
            combinations_r = itertools.combinations(all_selected_channels, r)
            channel_combinations.extend(combinations_r)

        th_ix = results.thresholds.index(threshold_metric)
        metrics = getattr(results, metric)
        metric_per_fold = [metrics[fold_i][th_ix] for fold_i in range(len(metrics))]

        base_metrics = getattr(base_results, metric)
        base_metric_per_fold = [base_metrics[fold_i][th_ix] for fold_i in range(len(base_metrics))]

        difference_per_fold = [metric_per_fold[fold_i] - base_metric_per_fold[fold_i]
                               for fold_i in range(len(metric_per_fold))]

        score_combinations = {}
        for combination in channel_combinations:
            combination_set = set(combination)
            folds_with_combination = [fold_i for fold_i in range(config.nb_folds)
                                     if combination_set.issubset(set(selected_channels[fold_i]))]
            if len(folds_with_combination) / config.nb_folds >= minimal_support:
                scores_with_combination = [difference_per_fold[fold_i] for fold_i in folds_with_combination]
                if not scores_with_combination:
                    score_with_combination = 0
                else:
                    score_with_combination = np.mean(scores_with_combination)
                scores_without_combination = [difference_per_fold[fold_i] for fold_i in range(config.nb_folds)
                                                if fold_i not in folds_with_combination]
                if not scores_without_combination:
                    score_without_combination = 0
                else:
                    score_without_combination = np.mean(scores_without_combination)
                score_combinations[combination] = score_with_combination - score_without_combination

        sorted_score_combinations = sorted(score_combinations.items(), key=lambda x: x[1], reverse=True)
        print(f"Differences per fold for {config.get_name()}: {difference_per_fold}")
        print(f"Channel combinations for {config.get_name()}:")
        for combination, score in sorted_score_combinations:
            print(f"Channels: {combination}, Score difference: {score}")

